import random
import openpyxl
import os

# 学生姓名列表
student_names = ['关羽', '张飞', '赵云', '马超', '黄忠']
# 为每位学生随机生成3门课程的成绩
scores = [[random.randrange(50, 101) for _ in range(3)] for _ in range(5)]
# 创建工作簿对象（Workbook）
wb = openpyxl.Workbook()
# 获取默认的工作表对象
sheet = wb.active
# 将工作表命名为“一年级二班”
sheet.title = '一年级二班'

# 定义表头
titles = ('姓名', '语文', '数学', '英语', '总分', '平均分')
# 添加表头数据到第一行
for index, title in enumerate(titles):
    sheet.cell(row=1, column=index + 1, value=title)
    # 为表头单元格设置粗体样式
    sheet.cell(row=1, column=index + 1).font = openpyxl.styles.Font(bold=True)

# 将学生姓名和考试成绩写入单元格
for row_index, student_name in enumerate(student_names):
    # 写入学生姓名到第一列
    sheet.cell(row=row_index + 2, column=1, value=student_name)

    # 写入三门课程的成绩
    for col_index, score in enumerate(scores[row_index]):
        sheet.cell(row=row_index + 2, column=col_index + 2, value=score)

# 在Excel中添加每位学生的总分和平均分
for row_index, student_scores in enumerate(scores):
    total_score = sum(student_scores)
    average_score = total_score / len(student_scores)

    # 写入总分到第五列
    sheet.cell(row=row_index + 2, column=5, value=total_score)

    # 写入平均分到第六列
    sheet.cell(row=row_index + 2, column=6, value=average_score)

# 保存Excel工作簿到指定路径
# 这里将文件保存在工作区的根目录
file_path = os.path.join('..', '考试成绩表.xlsx')
wb.save(file_path)

print(f"Excel文件已成功保存到: {file_path}")